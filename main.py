import tkinter as tk
from tkinter import ttk, simpledialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os

file = os.path.abspath("C:\\Users\\Lucas\\Documents\\Academia\\Academia.xlsx")

if os.path.exists(file):
    wb = load_workbook(file)
    sheet = wb.active
else:
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = 'Exercicio'
    sheet["B1"] = 'Carga'
    sheet["C1"] = 'Repetições'
    sheet["D1"] = 'Km'
    sheet["E1"] = 'Tempo'
    sheet["F1"] = 'Data'

    header_style = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid") 
    font_style = Font(bold=True)

    for col_num in range(1, 7):
        sheet.cell(row=1, column=col_num).fill = header_style
        sheet.cell(row=1, column=col_num).font = font_style

def inserirDados():
    exercicio = simpledialog.askstring("Exercício", "Digite o exercício")
    carga = simpledialog.askinteger("Carga", "Digite a carga")
    repeticoes = simpledialog.askinteger("Repetições", "Digite a quantidade de repetições")
    km = simpledialog.askinteger("Km", "Digite a quantidade de km")
    tempo = simpledialog.askinteger("Tempo", "Digite o tempo gasto (em minutos)")
    data = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    nova_linha = (exercicio, carga, repeticoes, km, tempo, data)
    sheet.append(nova_linha)

    print(f'Exercício: {exercicio}, Carga: {carga}, Repetições: {repeticoes}, Km: {km}, Tempo: {tempo}, Data: {data}')
    refreshTable()

def deletarLinha():
    selected_item = tree.selection()

    if not selected_item:
        messagebox.showinfo("Atenção", "Selecione uma linha para deletar.")
        return

    resposta = messagebox.askyesno("Deletar Linha", "Deseja realmente deletar esta linha?")

    if resposta:
        for item in selected_item:
            index = tree.index(item)
            sheet.delete_rows(index + 1)

        refreshTable()

def refreshTable():
    tree.delete(*tree.get_children())  # Limpa todas as linhas do Treeview

    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)

    wb.save(file)

root = tk.Tk()
root.title("Acompanhamento")

# Configurar Treeview com colunas e cabeçalhos
tree = ttk.Treeview(root, columns=("Exercicio", "Carga", "Repetições", "Km", "Tempo", "Data"), show="headings")
tree.heading("Exercicio", text="Exercicio")
tree.heading("Carga", text="Carga")
tree.heading("Repetições", text="Repetições")
tree.heading("Km", text="Km")
tree.heading("Tempo", text="Tempo (min)")
tree.heading("Data", text="Data")

for row in sheet.iter_rows(min_row=2, values_only=True):
    tree.insert("", "end", values=row)

tree.pack(expand=True, fill="both")

inserir_button = tk.Button(root, text="Inserir Dados", command=inserirDados)
inserir_button.pack()

delete_button = tk.Button(root, text="Deletar Linha", command=deletarLinha)
delete_button.pack()

refresh_button = tk.Button(root, text="Refresh", command=refreshTable)
refresh_button.pack()

root.mainloop()
